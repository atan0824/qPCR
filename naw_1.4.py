import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import tkinter
from tkinter import filedialog
import itertools
from openpyxl.styles import Font, Color, Alignment, Border, Side

print('Please select the qPCR raw data')

bold_font = Font(bold=True)
main_win = tkinter.Tk()
main_win.withdraw()
main_win.overrideredirect(True)
main_win.geometry('0x0+0+0')
main_win.deiconify()
main_win.lift()
main_win.focus_force()
main_win.sourceFile = filedialog.askopenfilename(parent=main_win, initialdir= "/",
title='Please select a directory')
main_win.destroy()
#print(main_win.sourceFile )
workbook = load_workbook(main_win.sourceFile)
sheet=workbook.active
count=9
ckdict={}
adict={}
bdict={}
samplelist=[]
sampledictlist=[]
jposcount=9


while True:
    jpos="B"+str(jposcount)
    if jposcount>120:
        break
    elif sheet[jpos].value==None:
        jposcount=jposcount+1
        continue
    elif len(sheet[jpos].value)==0:
        jposcount=jposcount+1
        continue
    elif str(sheet[jpos].value).lower() not in samplelist:
        samplelist.append(str(sheet[jpos].value).lower())
        jposcount=jposcount+1
        continue
    else:
        jposcount=jposcount+1
        continue


#print('yo')
#print(samplelist)
k=1
for a in samplelist:
    clone=str(a)
#    print(clone)
    a={}
    a[clone]=k
    k=k+1
    while True:
        pos="C"+str(count)
        opos="B"+str(count)
        lalapos='A'+str(count)
        if sheet[lalapos].value != None:
            if sheet[opos].value != None:
                if str(sheet[opos].value).lower() == clone:
                    if sheet[pos].value!=None:
                        if str(sheet[pos].value).lower() not in a:
                            if len(sheet[pos].value) != 0:
                                newpos="H"+str(count)
                                a.update({str(sheet[pos].value).lower():sheet[newpos].value})
                                count=count+1
                                continue
                            else:
                                count=count+1
                                continue
                        else:
                            count=count+1
                            continue
                    else:
                        count=count+1
                        continue
                else:
                    count=count+1
                    continue
            else:
                count=count+1
                continue
        else:
            sampledictlist.append(a)
            count=9
            break


#print(sampledictlist)




while True:
    try:
        CK=input("Housekeeping gene? ").lower()
        deltadictlist=[]
        for a in sampledictlist:
            deltadict={}
            for key, value in a.items():
                if key in samplelist:
                    deltadict[key]=value
                else:
                    delta=value-a[CK]
                    deltadict[key]=delta
            deltadictlist.append(deltadict)
        break
    except:
        print("Invalid internal control, not found in excel file provided please try again")
        continue
#print(deltadictlist)

while True:
    try:
        CKx=input("control treatment? ").lower()
        for a in deltadictlist:
            if CKx in a:
                j=a
                break
            else:
                continue
        j=='lol'
        break
    except:
        print("Invalid control treatment, not found in excel file provided please try again")
        continue



doubledeltadict={}
doubledeltadictlist=[]
for a in deltadictlist:
    doubledeltadict={}
    if a==j:
        continue
    for key,value in a.items():
        if key in samplelist:
            doubledeltadict[key]=0
            continue
        elif key==CK:
            continue
        else:
            doubledeltadict[key]=value-j[key]
    doubledeltadictlist.append(doubledeltadict)
alpha=['A','D','G','J','M','P','S','V','Y','AB','AE','AH','AK','AN','AQ','AT','AW','AZ']
beta=['B','E','H','K','N','Q','T','W','Z','AC','AF','AI','AL','AO','AR','AU','AX','BA']
omega=['C','F','I','L','O','R','U','X','AA','AD','AG','AJ','AM','AP','AS','AV','AY','BB']
#print(doubledeltadictlist)

foldchangedict={}
foldchangedictlist=[]
for b in doubledeltadictlist:
    foldchangedict={}
    for key,value in b.items():
        if key in samplelist:
            foldchangedict[key]=0
            continue
        else:
            newvalue=value*(-1)
            foldchangedict[key]=2**newvalue
    foldchangedictlist.append(foldchangedict)
#print(foldchangedictlist)



workbook = Workbook()
sheet = workbook.active
i=0
for a in foldchangedictlist:
    lol=list(a.keys())
    if CKx in lol:
        continue
    spos=str(alpha[i]+'2')
    kpos=str(alpha[i]+'1')
    sheet[kpos]="Sample Name"
    sheet[spos]=lol[0]
    sheet[kpos].font=bold_font
    i=i+1

pacount=4
for a,b,c in zip(foldchangedictlist,alpha,beta):
    pacount=4
    newpapos=str(b+'3')
    sheet[newpapos]='Genes'
    sheet[newpapos].font=bold_font
    nupapos=str(c+'3')
    sheet[nupapos]='Fold Change'
    sheet[nupapos].font=bold_font
    for gene in a.keys():
        keylist=a.keys()
        if gene in samplelist:
            continue
        samplepos=str(b+'2')
        if sheet[samplepos].value in keylist:
            papos=str(b+str(pacount))
            sheet[papos]=gene
            pacount=pacount+1

pcount=4
for a,b,c in zip(foldchangedictlist,beta,alpha):
    pcount=4
    for gene in a.keys():
        keylist=a.keys()
        if gene in samplelist:
            continue
        samplepos=str(c+'2')
        if sheet[samplepos].value in keylist:
            popos=str(b+str(pcount))
            sheet[popos]=a[gene]
            pcount=pcount+1

print('Task Successful')
koko=input("Save Filename? ")
koko=str(koko+'.xlsx')
workbook.save(filename=koko)
